import os
import sqlite3
import regions_and_cities
import openpyxl


# Функция преобразовывает столбцы таблицы users region_id и city_id из наименования в id
def parse_values_into_id(working_dict, list_of_values):
    for k, v in working_dict.items():
        for _dict in list_of_values:
            for key, value in _dict.items():
                if v == value:
                    _dict[key] = k
    return list_of_values


def excel_parse_file(file):
    try:
        title_dict = {}
        users_dict = {}
        users_list = []
        alphabet_dict = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        book = openpyxl.load_workbook(file)
        active_sheet = book.active
        starting_row = 3
        ending_row = active_sheet.max_row

        # Получаем названия столбцов
        for title in range(2, active_sheet.max_column + 1):
            title_dict[active_sheet.cell(row=2, column=title).value] = ''

        # Заполняем словарь значениями из колонок
        for row in range(starting_row, ending_row + 1):
            for i in range(len(alphabet_dict)):
                users_dict[i] = active_sheet[f'{alphabet_dict[i]}{row}'].value
            dict_copy = users_dict.copy()
            users_list.append(dict_copy)

        # Парсим данные в формат [{}], то есть массив со множеством словарей
        final_list = []
        for _dict in users_list:
            zipping_keys_dict = dict(zip(list(title_dict.keys()), list(_dict.keys())))
            zipping_keys_with_values = dict(zip(list(zipping_keys_dict.keys()), list(_dict.values())))
            final_list.append(zipping_keys_with_values)

        # Конвертируем регионы в id номера
        upd_regions = parse_values_into_id(regions_and_cities.regions_dict, final_list)
        # Конвертируем города в id номера
        upd_regions_and_cities = parse_values_into_id(regions_and_cities.cities_dict, upd_regions)

        return upd_regions_and_cities

    except FileNotFoundError:
        print(f'Файла ({file}) пока не существует :(')


def insert_into_users(data):
    try:
        list_of_tuples = []

        # Конвертируем данные в формат [()], то есть массив кортежей
        for _dict in data:
            temp_tuple = ()
            for v in _dict.values():
                temp_tuple += (v,)
            list_of_tuples.append(temp_tuple)

        if os.path.exists('./task.sqlite'):
            connect_to_db = sqlite3.connect('task.sqlite')
            query = """
                INSERT INTO users (second_name, first_name, patronymic, region_id, city_id, phone, email)
                VALUES (?,?,?,?,?,?,?);
           """
            with connect_to_db as db:
                cursor = db.cursor()
                for item in list_of_tuples:
                    cursor.execute(query, item)
        else:
            print('Файла БД еще не сформировано. Запустите сначала файл create_db.py')
    except TypeError:
        print('Проверьте директорию на наличие файла')


parse_file = excel_parse_file('test.xlsx')
insert_into_users(parse_file)
